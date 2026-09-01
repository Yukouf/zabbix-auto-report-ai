#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zabbix Auto Report - Rapport hebdomadaire Excel avec recommandations automatiques
Auteur : Youssef Guerniou (github.com/Yukouf)

Génère chaque semaine un rapport Excel multi-onglets à partir de l'API Zabbix :
alertes catégorisées, inventaire des hôtes, suivi de récurrence (nouvelles /
persistantes / résolues), tendance hebdomadaire, et recommandations de
diagnostic produites par un référentiel déterministe, complété par une IA
locale (Ollama) pour les cas inconnus. Aucune donnée ne quitte le serveur.

Configuration : via variables d'environnement ou fichier .env (voir .env.example).
Dépendance unique : openpyxl. Tout le reste est en bibliothèque standard Python.
"""

import json, argparse, urllib.request, ssl, smtplib, os, sys, re
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ------------------------------------------------------------------------------
# CONFIGURATION
# Toutes les valeurs sensibles sont lues depuis les variables d'environnement
# ou depuis un fichier .env placé à côté du script (voir .env.example).
# Ne JAMAIS écrire d'identifiants en dur dans ce fichier.
# ------------------------------------------------------------------------------
def _load_env(path=None):
    """Charge un fichier .env minimaliste (clé=valeur), sans dépendance externe."""
    path = path or os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
    except FileNotFoundError:
        pass

_load_env()

def _env(name, default=""):
    return os.environ.get(name, default).strip()

ZABBIX_URL = _env("ZABBIX_URL", "https://zabbix.example.local/api_jsonrpc.php")
# Normalisation : l'utilisateur peut donner "https://host" ou "https://host/" ;
# on complète avec le chemin API Zabbix si absent, pour éviter une erreur 404.
if not ZABBIX_URL.rstrip("/").endswith("/api_jsonrpc.php"):
    ZABBIX_URL = ZABBIX_URL.rstrip("/") + "/api_jsonrpc.php"
ZABBIX_USER = _env("ZABBIX_USER", "rapport-auto")
ZABBIX_PASS = _env("ZABBIX_PASS")
VERIFY_SSL = _env("VERIFY_SSL", "false").lower() == "true"
SMTP_SERVER = _env("SMTP_SERVER")
SMTP_PORT = int(_env("SMTP_PORT", "465"))
SMTP_USER = _env("SMTP_USER")
SMTP_PASS = _env("SMTP_PASS")
SMTP_FROM = _env("SMTP_FROM", "Zabbix Alertes <noreply@example.com>")
SMTP_SENDER = _env("SMTP_SENDER", "noreply@example.com")
EMAIL_TO = [e.strip() for e in _env("EMAIL_TO").split(",") if e.strip()]
COMPANY_NAME = _env("COMPANY_NAME", "Supervision")
REPORT_DIR = _env("REPORT_DIR", os.path.join(os.path.dirname(os.path.abspath(__file__)), "rapports"))
STATE_FILE = os.path.join(REPORT_DIR, "recurrence_state.json")
OLLAMA_URL = _env("OLLAMA_URL", "http://127.0.0.1:11434/api/generate")
OLLAMA_MODEL = _env("OLLAMA_MODEL", "qwen2.5:7b")


def _http_url(url: str) -> str:
    """Garde anti-SSRF : n'accepte que http/https (bloque file://, gopher://...)."""
    if not url.lower().startswith(("http://", "https://")):
        raise ValueError(f"schéma non-HTTP refusé: {url[:40]}")
    return url

# Alertes ignorées dans le rapport (bruit connu) : personnalisable via .env
EXCLUDED_PATTERNS = [p for p in _env("EXCLUDED_PATTERNS", r"Ethernet has changed to lower speed;Operating system description has changed;GoogleUpdater;Number of installed packages has been changed").split(";") if p]
EXCLUDED_SEVERITIES = ["0", "1"]
# Mots-clés identifiant vos équipements réseau (noms d'hôtes) : personnalisable via .env
NETWORK_KEYWORDS = [k.strip().lower() for k in _env("NETWORK_KEYWORDS", "aruba,hp-2530,switch").split(",") if k.strip()]
NETWORK_AGENTS = ["2"]

# Libellés d'affichage des catégories internes (les clés restent sans accent dans le code)
CAT_LABELS = {"Serveur": "Serveur", "Reseau": "Réseau", "Poste": "Poste", "Peripherique": "Périphérique"}

# Contraintes OS injectées dans le prompt IA (déterministe, décidé côté Python)
OS_CONSTRAINTS = {
    "Windows": "CONTRAINTE OS : cet hote est WINDOWS. Donne UNIQUEMENT des commandes PowerShell ou cmd (Get-Service, net start, Get-PSDrive, Get-Counter, Get-WinEvent). N'utilise JAMAIS df, du, systemctl, free, ip, journalctl, ethtool.",
    "Linux": "CONTRAINTE OS : cet hote est LINUX. Donne UNIQUEMENT des commandes bash (systemctl, df, du, free, ip, ss, journalctl). N'utilise JAMAIS Get-Service, net start, wmic, Get-PSDrive ni aucune commande PowerShell.",
    "Reseau": "CONTRAINTE OS : equipement RESEAU (switch). Donne UNIQUEMENT des commandes CLI switch (show interface brief, show interface <port>, show lldp info remote). N'utilise JAMAIS df, systemctl, ethtool ni PowerShell.",
    "Inconnu": "OS non identifie avec certitude : reste sur un diagnostic generique et prudent, sans commande propre a un OS specifique.",
}

ssl_ctx = ssl.create_default_context()
if not VERIFY_SSL:
    # Certificat interne auto-signé : désactivable via VERIFY_SSL=true dans .env
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=ssl_ctx))

def zabbix_api(method, params, auth=None):
    payload = {"jsonrpc": "2.0", "method": method, "params": params, "id": 1}
    headers = {"Content-Type": "application/json-rpc"}
    # apiinfo.version DOIT être appelé sans auth (rejeté sinon par Zabbix >= 6.4)
    if auth is not None and method not in ("user.login", "apiinfo.version"):
        headers["Authorization"] = f"Bearer {auth}"
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(ZABBIX_URL, data=data, headers=headers)
    resp = opener.open(req, timeout=30)
    result = json.loads(resp.read().decode("utf-8"))
    if "error" in result:
        raise Exception(f"Erreur API Zabbix: {result['error']}")
    return result["result"]

def get_auth_token():
    return zabbix_api("user.login", {"username": ZABBIX_USER, "password": ZABBIX_PASS})

def get_api_version(auth=None):
    try:
        return zabbix_api("apiinfo.version", {}, auth)
    except Exception:
        return "6.0"

def _version_tuple(v):
    parts = []
    for x in str(v).split("."):
        m = re.match(r"\d+", x)
        parts.append(int(m.group()) if m else 0)
    while len(parts) < 3:
        parts.append(0)
    return tuple(parts[:3])

def host_groups(host):
    # API >= 6.2 renvoie "hostgroups", API < 6.2 renvoie "groups"
    return host.get("hostgroups") or host.get("groups") or []

def get_hosts(auth):
    ver = _version_tuple(get_api_version(auth))
    grp_select = "selectHostGroups" if ver >= (6, 2, 0) else "selectGroups"
    params = {
        "output": ["hostid", "host", "name", "status", "inventory_mode"],
        "selectInterfaces": ["ip", "type", "available"],
        grp_select: ["name"],
        "sortfield": "name",
    }
    return zabbix_api("host.get", params, auth)

def get_problems(auth):
    return zabbix_api("problem.get", {"output": ["eventid", "objectid", "name", "severity", "clock", "r_clock", "acknowledged"], "selectTags": "extend", "recent": True, "sortfield": "eventid", "sortorder": "DESC", "suppressed": False}, auth)

def get_triggers(auth, trigger_ids):
    if not trigger_ids: return []
    return zabbix_api("trigger.get", {"output": ["triggerid", "description", "priority"], "triggerids": trigger_ids, "selectHosts": ["host", "name"], "expandDescription": True}, auth)

def get_host_availability(auth):
    hosts = zabbix_api("host.get", {"output": ["hostid"], "selectInterfaces": ["available"], "filter": {"status": 0}}, auth)
    a = u = k = 0
    for h in hosts:
        if h.get("interfaces"):
            v = h["interfaces"][0].get("available", "0")
            if v == "1": a += 1
            elif v == "2": u += 1
            else: k += 1
        else: k += 1
    return {"total": len(hosts), "available": a, "unavailable": u, "unknown": k}

def severity_name(sev):
    return {"0": "Non classé", "1": "Information", "2": "Avertissement", "3": "Moyen", "4": "Haut", "5": "Désastre"}.get(str(sev), "Inconnu")

def inventory_mode_label(mode):
    return {"-1": "Désactivé", "0": "Manuel", "1": "Automatique"}.get(str(mode), "Inconnu")

def is_excluded(problem):
    if str(problem.get("severity", "0")) in EXCLUDED_SEVERITIES: return True
    name = problem.get("name", "") or ""
    for pattern in EXCLUDED_PATTERNS:
        if re.search(pattern, name, re.IGNORECASE): return True
    return False

# ---- Récurrence + historique (état persistant) --------------------------------
# Format du fichier d'état :
# { "alerts": { "<clé>": {"count": N, "week": "2026-W29", "last": "2026-07-15"} },
#   "history": [ {"week": "2026-W29", "date": "...", "alertes": N, "critiques": N} ] }
# Rétro-compatible avec l'ancien format plat (dict clé -> {count, last}).
def load_state():
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return {"alerts": {}, "history": []}
    if isinstance(raw, dict) and ("alerts" in raw or "history" in raw):
        return {"alerts": raw.get("alerts", {}), "history": raw.get("history", [])}
    return {"alerts": raw if isinstance(raw, dict) else {}, "history": []}

def save_state(alerts, history):
    try:
        os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"alerts": alerts, "history": history}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[WARN] Sauvegarde état récurrence: {e}")

def problem_key(host, problem):
    # Clé stable : on neutralise les valeurs variables (tailles, durées, %) pour
    # qu'une même alerte récurrente garde la même clé d'une semaine à l'autre.
    norm = re.sub(r"\d+", "#", problem or "")
    return f"{host}||{norm}".strip().lower()

def iso_week(dt):
    return dt.strftime("%G-W%V")

def week_of_datestr(datestr):
    try:
        return iso_week(datetime.strptime(datestr, "%Y-%m-%d"))
    except Exception:
        return ""
# -----------------------------------------------------------------------------

def detect_os(problem_name, host_name, hosts_data):
    """Détermination déterministe de l'OS. Ordre : préfixe alerte -> groupes -> catégorie réseau."""
    name = problem_name or ""
    if name.startswith("Windows:"): return "Windows"
    if name.startswith("Linux:"):   return "Linux"
    for h in hosts_data:
        if h.get("name", "") == host_name or h.get("host", "") == host_name:
            grp = " ".join(g.get("name", "").lower() for g in host_groups(h))
            if "windows" in grp: return "Windows"
            if "linux" in grp:   return "Linux"
    if classify_host(host_name, hosts_data) == "Reseau":
        return "Reseau"
    return "Inconnu"

def ask_ollama(prompt):
    try:
        payload = {"model": OLLAMA_MODEL, "prompt": prompt, "stream": False, "options": {"temperature": 0.3, "num_predict": 150}}
        req = urllib.request.Request(_http_url(OLLAMA_URL), data=json.dumps(payload).encode("utf-8"), headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=120) as resp:  # nosemgrep: python.lang.security.audit.dynamic-urllib-use-detected — URL validée en amont par _http_url
            return json.loads(resp.read().decode("utf-8")).get("response", "").strip()
    except Exception as e:
        return f"Erreur IA: {e}"

def get_recommendation(problem_name, host_name, severity, duration, os_hint="Inconnu"):
    system = """Tu es ingenieur SecOps. Tu analyses des alertes Zabbix sur une infra heterogene (serveurs Linux, serveurs Windows, switches Aruba HP-2530, postes de travail, imprimantes).
Pour chaque alerte, tu produis UNE recommandation en 2 phrases maximum, en francais correct AVEC les accents, directement actionnable par un admin systeme.
METHODE :
1. Le type d'hote (OS) t'est IMPOSE par la contrainte fournie : respecte-la a la lettre, ne le redevine pas.
2. Donne la premiere commande concrete de diagnostic, ADAPTEE a l'OS impose ET a la metrique de l'alerte. Texte brut uniquement : JAMAIS de backticks, JAMAIS de markdown.
3. Mentionne la piste de resolution si evidente.
CONTRAINTES STRICTES :
- Contexte = infrastructure datacenter professionnelle. JAMAIS de solutions grand public.
- Respecte STRICTEMENT la contrainte OS donnee a la fin du message.
- La commande doit cibler la METRIQUE exacte de l'alerte : alerte memoire -> compteurs memoire (PAS le CPU) ; alerte disque -> espace disque ; alerte service -> etat du service ; alerte CPU -> charge CPU.
- Alerte switch (Aruba/HP-2530/Link down) : commandes SSH switch (show interface brief, show interface 21, show lldp info remote). JAMAIS ethtool/ifconfig.
- Alerte imprimante ("IMP ", "Toner") : verification manuelle sur place.
- JAMAIS de commandes destructrices. Uniquement du diagnostic.
- Agent "not available" = probleme reseau/service/firewall, JAMAIS mise a jour de package.
- Francais correct avec accents. Style direct, imperatif. Reponse brute, sans preambule, sans conclusion, sans markdown."""
    user = (f"Hote: {host_name}\nProbleme: {problem_name}\nSeverite: {severity}\nDuree: {duration}\n\n"
            f"{OS_CONSTRAINTS.get(os_hint, OS_CONSTRAINTS['Inconnu'])}")
    payload = {"model": OLLAMA_MODEL, "system": system, "prompt": user, "stream": False, "options": {"temperature": 0.2, "num_predict": 180}}
    data = json.dumps(payload).encode("utf-8")
    for attempt in range(2):
        try:
            req = urllib.request.Request(_http_url(OLLAMA_URL), data=data, headers={"Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=150) as resp:  # nosemgrep: python.lang.security.audit.dynamic-urllib-use-detected — URL validée en amont par _http_url
                return json.loads(resp.read().decode("utf-8")).get("response", "").strip()
        except Exception as e:
            if attempt == 0: continue
            return f"Erreur IA: {e}"

# ---- Référentiel déterministe alerte -> commande --------------------------------
# Couvre les motifs d'alerte récurrents. Match sur le libellé (en minuscules),
# commande choisie selon l'OS déterministe. "*" = indépendant de l'OS.
# Premier motif qui matche ET qui couvre l'OS courant gagne ; sinon on continue,
# et à défaut on tombe sur l'IA (avec contrainte OS).
COMMAND_TABLE = [
    {   # Agent Zabbix indisponible (AVANT 'service' : réseau/firewall, jamais màj package)
        "patterns": ["agent is not available", "agent not available",
                     "agent is unavailable", "agent unavailable",
                     "is unreachable", "zabbix agent on"],
        "commands": {
            "Windows": "L'agent Zabbix ne répond plus (en général réseau ou service, pas une mise à jour).\n1) Vérifier que le service tourne :\nGet-Service 'Zabbix Agent*'\n2) Tester l'accès réseau au port de l'agent :\nTest-NetConnection <serveur-zabbix> -Port 10050\nPuis : si le test échoue, ouvrir le port 10050 dans le pare-feu de l'hôte.",
            "Linux":   "L'agent Zabbix ne répond plus (en général réseau ou service, pas une mise à jour).\n1) Vérifier que le service tourne :\nsystemctl status zabbix-agent2\n2) Vérifier que le port 10050 écoute :\nss -lntp | grep 10050\nPuis : si rien n'écoute, redémarrer l'agent ; sinon ouvrir le port 10050 (firewalld/nftables).",
        },
    },
    {   # Redémarrage récent (libellé "has been restarted (uptime < 10m)")
        "patterns": ["has been restarted", "uptime <"],
        "commands": {
            "Windows": "La machine vient de redémarrer.\n1) Voir l'heure du dernier démarrage :\nGet-CimInstance Win32_OperatingSystem | Select LastBootUpTime\n2) Identifier l'origine du redémarrage (events 1074/6008) :\nGet-WinEvent -FilterHashtable @{LogName='System'; Id=1074,6008} -MaxEvents 5\nPuis : si le redémarrage n'était pas planifié, chercher la cause (mise à jour, crash, coupure).",
            "Linux":   "La machine vient de redémarrer.\n1) Confirmer l'heure du démarrage :\nuptime -s\n2) Vérifier si le redémarrage était volontaire :\nlast -x reboot | head\n3) Regarder les erreurs du boot précédent :\njournalctl -b -1 -p err --no-pager | tail -20\nPuis : si non planifié, chercher la cause (OOM, kernel panic, coupure électrique).",
        },
    },
    {   # Espace disque
        "patterns": ["space is", "fs [", "disk space"],
        "commands": {
            "Windows": "Disque presque plein : trouver ce qui prend la place.\n1) Voir l'espace libre par disque :\nGet-PSDrive -PSProvider FileSystem\n2) Lister les 20 plus gros fichiers du disque concerné :\nGet-ChildItem <lecteur>:\\ -Recurse -EA SilentlyContinue | Sort Length -Desc | Select -First 20\nPuis : supprimer les logs/fichiers temp inutiles, ou agrandir le disque si ça revient.",
            "Linux":   "Disque presque plein : trouver ce qui prend la place.\n1) Voir l'espace libre par système de fichiers :\ndf -h\n2) Trouver les gros répertoires du point de montage concerné :\ndu -xh <point_montage> --max-depth=2 | sort -rh | head\nPuis : faire le ménage (logs) ou agrandir le volume si ça revient.",
        },
    },
    {   # Swap (Linux)
        "patterns": ["swap"],
        "commands": {
            "Linux": "Le swap est trop utilisé (la RAM sature et déborde sur le disque).\n1) Voir l'état mémoire et swap :\nfree -m\n2) Trouver le processus qui consomme la mémoire :\nps aux --sort=-%mem | head\nPuis : traiter le processus en cause avant d'envisager d'agrandir le swap.",
        },
    },
    {   # Mémoire / pages
        "patterns": ["memory pages", "pages/sec", "memory", "out of memory"],
        "commands": {
            "Windows": "Forte pression sur la mémoire.\n1) Mesurer la pagination et la mémoire disponible :\nGet-Counter '\\Memory\\Pages/sec','\\Memory\\Available MBytes'\n2) Trouver le processus le plus gourmand en mémoire :\nGet-Process | Sort WS -Desc | Select -First 10\nPuis : identifier/limiter ce processus, ou ajouter de la RAM si récurrent.",
            "Linux":   "Forte pression sur la mémoire.\n1) Voir l'état mémoire :\nfree -m\n2) Trouver le processus le plus gourmand :\nps aux --sort=-%mem | head\n3) Vérifier si le noyau a tué un processus (OOM) :\ndmesg -T | grep -i oom\nPuis : traiter le processus, ou ajouter de la RAM si récurrent.",
        },
    },
    {   # CPU / charge
        "patterns": ["cpu", "load average", "processor", "utilization too high"],
        "commands": {
            "Windows": "Charge CPU élevée.\n1) Mesurer l'utilisation CPU :\nGet-Counter '\\Processor(_Total)\\% Processor Time'\n2) Trouver le processus le plus gourmand :\nGet-Process | Sort CPU -Desc | Select -First 10\nPuis : identifier le processus et son origine (boucle, tâche planifiée...).",
            "Linux":   "Charge CPU élevée.\n1) Voir la charge moyenne :\nuptime\n2) Voir les processus actifs en temps réel :\ntop -b -n1 | head -20\nPuis : identifier le processus en haut de liste et son origine.",
        },
    },
    {   # Bande passante interface (AVANT le switch : cas conteneur/veth Linux)
        "patterns": ["bandwidth"],
        "commands": {
            "Linux":  "Trafic réseau très élevé sur une interface.\n1) Voir le trafic de l'interface :\nip -s link show <iface>\n2) Voir le débit en temps réel :\nsar -n DEV 1 5\nPuis : identifier le service/conteneur à l'origine du flux.",
            "Reseau": "Trafic élevé sur un port du switch.\n1) Voir l'état et le débit du port :\nshow interface <port>\n2) Identifier l'équipement connecté :\nshow lldp info remote-device <port>\nPuis : vérifier l'équipement en face.",
        },
    },
    {   # Lien / interface switch
        "patterns": ["link down", "lldp", "interface is down", "port is down"],
        "commands": {
            "Reseau": "Un port du switch est tombé.\n1) Voir l'état général des ports :\nshow interface brief\n2) Détailler le port concerné (erreurs, flaps) :\nshow interface <port>\n3) Identifier l'équipement déconnecté :\nshow lldp info remote-device <port>\nPuis : vérifier le câble et l'équipement en face.",
        },
    },
    {   # Service arrêté (après agent)
        "patterns": ["is not running", "not running", "service", "windefend"],
        "commands": {
            "Windows": "Un service Windows est arrêté alors qu'il devrait tourner.\n1) Vérifier son état :\nGet-Service '<nom_service>'\n2) Le redémarrer :\nStart-Service '<nom_service>'\n3) S'il retombe, regarder les erreurs système :\nGet-WinEvent -LogName System -MaxEvents 20\nPuis : traiter la cause si le service ne tient pas.",
            "Linux":   "Un service est arrêté alors qu'il devrait tourner.\n1) Vérifier son état :\nsystemctl status <service>\n2) Le redémarrer :\nsystemctl restart <service>\n3) En cas d'échec, lire les logs du service :\njournalctl -u <service> -n 50 --no-pager\nPuis : traiter la cause si le service ne tient pas.",
        },
    },
    {   # Synchro horaire / NTP
        "patterns": ["time is out of sync", "out of sync", "system time", "ntp"],
        "commands": {
            "Windows": "L'horloge de la machine n'est plus synchronisée.\n1) Voir l'état de synchro :\nw32tm /query /status\n2) Forcer une resynchronisation :\nw32tm /resync /force\nPuis : vérifier la source de temps (contrôleur de domaine / NTP).",
            "Linux":   "L'horloge de la machine n'est plus synchronisée.\n1) Voir l'état de synchro :\ntimedatectl status\n2) Forcer une resynchronisation :\nchronyc makestep\nPuis : si besoin, redémarrer le service : systemctl restart chronyd.",
        },
    },
    {   # Imprimante / consommable (couvre aussi les hôtes nommés IMP-xx)
        "patterns": ["toner", "imp-", "imp ", "printer", "cartridge", "drum"],
        "commands": {
            "*": "Problème de consommable sur une imprimante.\nVérification manuelle sur place : niveau du toner/tambour et état physique de l'imprimante.\nPuis : remplacer le consommable si nécessaire.",
        },
    },
]

def match_reco(problem_name, os_hint):
    """Retourne une recommandation déterministe si l'alerte matche un motif connu pour l'OS, sinon None (-> IA)."""
    text = (problem_name or "").lower()
    name = problem_name or ""
    def fill(cmd):
        # Remplace les placeholders par les valeurs réelles extraites du libellé de l'alerte
        svc = extract_service_name(name)
        if svc:
            cmd = cmd.replace("<nom_service>", svc).replace("<service>", svc)
        mdrv = re.search(r'\(([A-Za-z]):\)', name)          # disque Windows : (C:) ou Donnees(E:)
        if mdrv:
            cmd = cmd.replace("<lecteur>", mdrv.group(1))
        else:
            mmnt = re.search(r'[Ff][Ss] \[([^\]]+)\]', name)  # point de montage Linux : FS [/]
            if mmnt:
                cmd = cmd.replace("<point_montage>", mmnt.group(1))
        mif = re.search(r'[Ii]nterface ([^\s:]+)', name)    # interface Linux : Interface ztzlgow4uw
        if mif and not mif.group(1).isdigit():
            cmd = cmd.replace("<iface>", mif.group(1))
        mport = re.search(r'(?:interface|port)\s+(\d+)', name, re.IGNORECASE)  # port switch : interface 21
        if mport:
            cmd = cmd.replace("<port>", mport.group(1))
        return cmd
    for rule in COMMAND_TABLE:
        if any(pat in text for pat in rule["patterns"]):
            cmds = rule["commands"]
            if "*" in cmds:
                return fill(cmds["*"])
            if os_hint in cmds:
                return fill(cmds[os_hint])
            # motif connu mais OS non couvert : on laisse une autre règle ou l'IA décider
            continue
    return None

def extract_service_name(problem_name):
    """Extrait le nom court du service depuis le libellé Zabbix (entre guillemets). None si absent."""
    m = re.search(r'"([^"]+)"', problem_name or "")
    return m.group(1) if m else None
# ---------------------------------------------------------------------------------

def classify_host(host_name, hosts_data):
    name_lower = (host_name or "").lower()
    for kw in NETWORK_KEYWORDS:
        if kw in name_lower: return "Reseau"
    for h in hosts_data:
        if h.get("name", "") == host_name or h.get("host", "") == host_name:
            if h.get("interfaces"):
                if h["interfaces"][0].get("type") in NETWORK_AGENTS: return "Reseau"
    if name_lower.startswith("pc") or name_lower.startswith("hpmx"):
        return "Poste"
    if name_lower.startswith("imp-") or name_lower.startswith("imp "):
        return "Peripherique"
    return "Serveur"

def generate_report(hosts, problems, triggers_map, availability):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    now = datetime.now()
    now_week = iso_week(now)
    wk_s = now - timedelta(days=now.weekday())
    wk_e = wk_s + timedelta(days=6)
    thin_border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    bottom_border = Border(bottom=Side(style='medium', color='1F4E79'))
    title_font = Font(name='Calibri', bold=True, size=18, color='1F4E79')
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='808080')
    section_font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    normal_font = Font(name='Calibri', size=10)
    metric_label_font = Font(name='Calibri', size=9, color='606060')
    reco_font = Font(name='Calibri', size=9, color='1565C0')
    light_gray = PatternFill('solid', fgColor='F8F9FA')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    green_metric = PatternFill('solid', fgColor='E8F5E9')
    red_metric = PatternFill('solid', fgColor='FFEBEE')
    yellow_metric = PatternFill('solid', fgColor='FFF8E1')
    blue_metric = PatternFill('solid', fgColor='E3F2FD')
    alert_bg = PatternFill('solid', fgColor='FFF0F0')
    dark_header = PatternFill('solid', fgColor='1F4E79')
    sev_fills = {"Désastre": PatternFill('solid', fgColor='D32F2F'), "Haut": PatternFill('solid', fgColor='E64A19'), "Moyen": PatternFill('solid', fgColor='F57C00'), "Avertissement": PatternFill('solid', fgColor='FBC02D')}
    sev_fonts = {"Désastre": Font(name='Calibri', size=10, bold=True, color='FFFFFF'), "Haut": Font(name='Calibri', size=10, bold=True, color='FFFFFF'), "Moyen": Font(name='Calibri', size=10, bold=True, color='FFFFFF'), "Avertissement": Font(name='Calibri', size=10, bold=True, color='333333')}
    cat_configs = {
        "Serveur": {"fill": PatternFill('solid', fgColor='1565C0'), "icon": "SERVEURS", "row_fill": PatternFill('solid', fgColor='F5F9FF')},
        "Reseau": {"fill": PatternFill('solid', fgColor='E65100'), "icon": "ÉQUIPEMENTS RÉSEAU", "row_fill": PatternFill('solid', fgColor='FFF8F0')},
        "Poste": {"fill": PatternFill('solid', fgColor='2E7D32'), "icon": "POSTES DE TRAVAIL", "row_fill": PatternFill('solid', fgColor='F5FFF5')},
        "Peripherique": {"fill": PatternFill('solid', fgColor='6A1B9A'), "icon": "PÉRIPHÉRIQUES", "row_fill": PatternFill('solid', fgColor='FCF5FF')},
    }
    filtered = [p for p in problems if not is_excluded(p)]
    excluded_count = len(problems) - len(filtered)
    categorized = {"Serveur": [], "Reseau": [], "Poste": [], "Peripherique": []}
    total = len(filtered)

    # Récurrence : on charge l'état du rapport précédent, on reconstruit l'état courant.
    state = load_state()
    prev_alerts = state.get("alerts", {})
    history = state.get("history", [])
    new_state = {}
    n_table = n_ia = 0

    for idx, p in enumerate(filtered):
        ti = triggers_map.get(p.get("objectid", ""), {})
        hosts_list = ti.get("hosts") or []
        hn = hosts_list[0].get("name", "") if hosts_list else ""
        cat = classify_host(hn, hosts)
        if cat not in categorized: cat = "Serveur"
        try: et = datetime.fromtimestamp(int(p.get("clock", 0)))
        except: et = now
        dur = now - et
        total_sec = int(dur.total_seconds())
        d = total_sec // 86400; h = (total_sec % 86400) // 3600; m = (total_sec % 3600) // 60
        duration_str = f"{d}j {h}h" if d > 0 else f"{h}h {m}m"
        sev = severity_name(p.get("severity", "0"))

        # Compteur de récurrence : +1 UNE SEULE FOIS PAR SEMAINE ISO.
        # Plusieurs exécutions la même semaine conservent la même valeur.
        key = problem_key(hn, p.get("name", ""))
        if key in new_state:
            weeks = new_state[key]["count"]
        else:
            prev = prev_alerts.get(key)
            if prev:
                prev_week = prev.get("week") or week_of_datestr(prev.get("last", ""))
                if prev_week == now_week:
                    weeks = max(1, int(prev.get("count", 1)))   # même semaine : pas d'incrément
                else:
                    weeks = int(prev.get("count", 0)) + 1        # nouvelle semaine : +1
            else:
                weeks = 1
            new_state[key] = {"count": weeks, "week": now_week, "last": now.strftime("%Y-%m-%d")}
        # Première apparition = étiquette explicite "NOUVELLE" (demande de Pierre)
        seen_str = "NOUVELLE" if weeks == 1 else f"{weeks} sem."

        os_hint = detect_os(p.get("name", ""), hn, hosts)
        # Référentiel déterministe d'abord ; IA locale uniquement si motif inconnu
        reco = match_reco(p.get("name", ""), os_hint)
        source = "Référentiel" if reco else "IA"
        print(f"  [{idx+1}/{total}] {hn} [{os_hint}/{source}]: {p.get('name','')[:42]}...", end=" ", flush=True)
        if not reco:
            reco = get_recommendation(p.get("name", ""), hn, sev, duration_str, os_hint)
            n_ia += 1
            # L'IA peut renvoyer des sauts de ligne / markdown parasites : on nettoie
            reco = reco.replace("\n", " ").replace("```", "").replace("`", "").strip()
            if len(reco) > 300: reco = reco[:297] + "..."
        else:
            n_table += 1
            # Référentiel : on conserve la mise en forme en étapes (sauts de ligne volontaires)
            reco = reco.strip()
        print("OK")
        categorized[cat].append({"date": et.strftime("%d/%m/%Y %H:%M"), "host": hn, "severity": sev, "problem": p.get("name", ""), "duration": duration_str, "seen": seen_str, "source": source, "ack": "Oui" if str(p.get("acknowledged", "0")) == "1" else "Non", "sev_code": str(p.get("severity", "0")), "reco": reco})

    # Évolution : mêmes règles que la colonne "Vu depuis".
    # Nouvelle = vue pour la première fois cette semaine (count == 1),
    # même si le script a déjà tourné plusieurs fois dans la semaine.
    prev_keys = set(prev_alerts.keys())
    cur_keys = set(new_state.keys())
    new_keys = {k for k, v in new_state.items() if int(v.get("count", 1)) == 1}
    persistent_keys = cur_keys - new_keys
    resolved_keys = prev_keys - cur_keys
    resolved_list = []
    for k in sorted(resolved_keys, key=lambda x: -int(prev_alerts[x].get("count", 0))):
        hn_r, patt = (k.split("||", 1) + [""])[:2]
        resolved_list.append({"host": hn_r.upper(), "pattern": patt, "count": prev_alerts[k].get("count", "?")})

    print(f"[OK] Recommandations : {n_table} via référentiel déterministe, {n_ia} via IA locale")
    print(f"[OK] Évolution : {len(new_keys)} nouvelles / {len(persistent_keys)} persistantes / {len(resolved_keys)} résolues")

    sev_count = {"Désastre": 0, "Haut": 0, "Moyen": 0, "Avertissement": 0}
    for cat_list in categorized.values():
        for p in cat_list:
            if p["severity"] in sev_count: sev_count[p["severity"]] += 1

    # Historique hebdo (tendance) : une entrée par semaine ISO, mise à jour si re-run
    hist_entry = {"week": now_week, "date": now.strftime("%Y-%m-%d"), "alertes": len(filtered), "critiques": sev_count['Haut'] + sev_count['Désastre']}
    if history and history[-1].get("week") == now_week:
        history[-1] = hist_entry
    else:
        history.append(hist_entry)
    history = history[-12:]
    save_state(new_state, history)

    ws = wb.active; ws.title = "Rapport Hebdo"; ws.sheet_properties.tabColor = "1F4E79"
    for col, w in {'A': 14, 'B': 15, 'C': 12, 'D': 33, 'E': 9, 'F': 10, 'G': 6, 'H': 60}.items(): ws.column_dimensions[col].width = w
    ws.merge_cells('A2:H2'); ws['A2'] = 'RAPPORT DE SUPERVISION'; ws['A2'].font = title_font; ws['A2'].alignment = Alignment(horizontal='left', vertical='center'); ws.row_dimensions[2].height = 35
    ws.merge_cells('A3:H3'); ws['A3'] = f'Semaine du {wk_s.strftime("%d/%m/%Y")} au {wk_e.strftime("%d/%m/%Y")}  |  {COMPANY_NAME}  |  {now.strftime("%d/%m/%Y %H:%M")}  |  Recommandations : référentiel déterministe + IA locale Qwen2.5 (voir onglet Fonctionnement)'; ws['A3'].font = subtitle_font
    for c in range(1, 9): ws.cell(row=4, column=c).border = bottom_border
    row = 6
    metrics = [(availability['total'], "Hôtes total", blue_metric, '1565C0'), (availability['available'], "Disponibles", green_metric, '2E7D32'), (availability['unavailable'], "Non disponibles", red_metric, 'C62828'), (len(filtered), "Alertes", yellow_metric, 'E65100'), (sev_count['Haut'] + sev_count['Désastre'], "Critiques", red_metric if sev_count['Haut'] + sev_count['Désastre'] > 0 else green_metric, 'C62828' if sev_count['Haut'] + sev_count['Désastre'] > 0 else '2E7D32'), (len(new_keys), "Nouvelles", yellow_metric if new_keys else green_metric, 'E65100' if new_keys else '2E7D32'), (excluded_count, "Filtrées", light_gray, '808080')]
    for i, (val, label, fill, color) in enumerate(metrics):
        col = i + 1; cell_v = ws.cell(row=row, column=col, value=val); cell_v.font = Font(name='Calibri', size=20, bold=True, color=color); cell_v.fill = fill; cell_v.alignment = Alignment(horizontal='center', vertical='center'); cell_v.border = thin_border; ws.row_dimensions[row].height = 40
        cell_l = ws.cell(row=row+1, column=col, value=label); cell_l.font = metric_label_font; cell_l.fill = fill; cell_l.alignment = Alignment(horizontal='center', vertical='center'); cell_l.border = thin_border
    row = 9
    # Ligne d'évolution vs rapport précédent
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = f"  ÉVOLUTION : {len(new_keys)} nouvelle(s) cette semaine  |  {len(persistent_keys)} persistante(s)  |  {len(resolved_keys)} résolue(s) depuis le dernier rapport (détail onglet Analyse)"
    ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='1565C0'); ws[f'A{row}'].fill = blue_metric; ws[f'A{row}'].border = thin_border
    row += 2
    attention = [f"{p['host']}: {p['problem'][:50]}" for cl in categorized.values() for p in cl if p["severity"] in ("Haut", "Désastre")]
    if attention:
        ws.merge_cells(f'A{row}:H{row}'); ws[f'A{row}'] = "POINTS D'ATTENTION"; ws[f'A{row}'].font = Font(name='Calibri', bold=True, size=10, color='C62828'); ws[f'A{row}'].fill = alert_bg; ws[f'A{row}'].border = thin_border
        for item in attention[:5]:
            row += 1; ws.merge_cells(f'A{row}:H{row}'); ws[f'A{row}'] = f'  {item}'; ws[f'A{row}'].font = Font(name='Calibri', size=10, color='C62828'); ws[f'A{row}'].fill = alert_bg; ws[f'A{row}'].border = thin_border
    row += 2; hdrs = ['Date', 'Hôte', 'Sévérité', 'Problème', 'Durée', 'Vu depuis', 'Ack', 'Recommandation']
    for cat_key in ["Serveur", "Reseau", "Poste", "Peripherique"]:
        items = categorized.get(cat_key, [])
        if not items: continue
        conf = cat_configs[cat_key]
        ws.merge_cells(f'A{row}:H{row}'); ws[f'A{row}'] = f'  {conf["icon"]}  ({len(items)})'; ws[f'A{row}'].font = section_font; ws[f'A{row}'].fill = conf["fill"]; ws[f'A{row}'].alignment = Alignment(vertical='center'); ws[f'A{row}'].border = thin_border; ws.row_dimensions[row].height = 28; row += 1
        for ci, hdr in enumerate(hdrs, 1):
            cell = ws.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='37474F'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
        ws.row_dimensions[row].height = 22; row += 1
        items.sort(key=lambda x: {"Désastre": 0, "Haut": 1, "Moyen": 2, "Avertissement": 3}.get(x["severity"], 5))
        for idx2, prob in enumerate(items):
            row_fill = conf["row_fill"] if idx2 % 2 == 0 else white_fill
            for ci, v in enumerate([prob["date"], prob["host"], prob["severity"], prob["problem"], prob["duration"], prob["seen"], prob["ack"], prob["reco"]], 1):
                cell = ws.cell(row=row, column=ci, value=v); cell.font = reco_font if ci == 8 else normal_font; cell.border = thin_border; cell.fill = row_fill
                if ci == 8:
                    cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='center' if ci in (6, 7) else 'general')
            sev = prob["severity"]
            if sev in sev_fills: ws.cell(row=row, column=3).fill = sev_fills[sev]; ws.cell(row=row, column=3).font = sev_fonts.get(sev, normal_font); ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='top')
            if prob["ack"] == "Oui":
                ac = ws.cell(row=row, column=7); ac.font = Font(name='Calibri', size=10, bold=True, color='2E7D32')
            if prob["seen"] == "NOUVELLE":
                nc = ws.cell(row=row, column=6); nc.font = Font(name='Calibri', size=9, bold=True, color='E65100')
            # Hauteur adaptée au nombre de lignes de la reco (col. H, largeur ~60)
            nb_lines = sum(max(1, -(-len(seg) // 58)) for seg in str(prob["reco"]).split("\n"))
            ws.row_dimensions[row].height = min(260, max(55, nb_lines * 13 + 6)); row += 1
        row += 1

    # Feuille 2 : Inventaire
    ws2 = wb.create_sheet("Inventaire Hôtes"); ws2.sheet_properties.tabColor = "2E7D32"
    for col, w in {'A': 22, 'B': 16, 'C': 8, 'D': 28, 'E': 12, 'F': 16, 'G': 14, 'H': 14}.items(): ws2.column_dimensions[col].width = w
    ws2.merge_cells('A1:H1'); ws2['A1'] = 'INVENTAIRE DES HÔTES'; ws2['A1'].font = title_font; ws2.row_dimensions[1].height = 35
    for c in range(1, 9): ws2.cell(row=2, column=c).border = bottom_border
    for ci, hdr in enumerate(['Nom', 'Adresse IP', 'Agent', 'Groupes', 'État', 'Disponibilité', 'Inventaire', 'Type'], 1):
        cell = ws2.cell(row=3, column=ci, value=hdr); cell.font = header_font; cell.fill = dark_header; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
    ws2.row_dimensions[3].height = 24
    for ri, host in enumerate(hosts, 4):
        ip = at = av = ""
        if host.get("interfaces"):
            ifc = host["interfaces"][0]; ip = ifc.get("ip", ""); at = {"1": "ZBX", "2": "SNMP", "3": "IPMI", "4": "JMX"}.get(ifc.get("type", "1"), "?"); av = {"0": "Inconnu", "1": "Disponible", "2": "Non disponible"}.get(ifc.get("available", "0"), "?")
        gr = ", ".join([g.get("name", "") for g in host_groups(host)]); st = "Activé" if str(host.get("status", "0")) == "0" else "Désactivé"; inv = inventory_mode_label(host.get("inventory_mode", "-1")); cat = classify_host(host.get("name", ""), hosts)
        row_fill = light_gray if (ri % 2 == 0) else white_fill
        for ci, v in enumerate([host.get("name", ""), ip, at, gr, st, av, inv, CAT_LABELS.get(cat, cat)], 1):
            cell = ws2.cell(row=ri, column=ci, value=v); cell.font = normal_font; cell.border = thin_border; cell.fill = row_fill
        if av == "Non disponible":
            for c in range(1, 9): ws2.cell(row=ri, column=c).fill = PatternFill('solid', fgColor='FFEBEE'); ws2.cell(row=ri, column=c).font = Font(name='Calibri', size=10, color='C62828')
        elif st == "Désactivé":
            for c in range(1, 9): ws2.cell(row=ri, column=c).fill = light_gray; ws2.cell(row=ri, column=c).font = Font(name='Calibri', size=10, color='999999')
        # Inventaire désactivé : on grise discrètement la cellule pour expliquer l'écart avec hostinventories.php
        if inv == "Désactivé":
            ic = ws2.cell(row=ri, column=7); ic.font = Font(name='Calibri', size=9, italic=True, color='B0B0B0'); ic.alignment = Alignment(horizontal='center')
        else:
            ws2.cell(row=ri, column=7).alignment = Alignment(horizontal='center')
        if cat in cat_configs: ws2.cell(row=ri, column=8).fill = cat_configs[cat]["fill"]; ws2.cell(row=ri, column=8).font = Font(name='Calibri', size=9, bold=True, color='FFFFFF'); ws2.cell(row=ri, column=8).alignment = Alignment(horizontal='center')
    # Note de bas de tableau : explique le 27 vs 25 (page hostinventories.php)
    note_row = len(hosts) + 5
    ws2.merge_cells(f'A{note_row}:H{note_row}')
    ws2[f'A{note_row}'] = "Note : les hôtes dont l'inventaire est \"Désactivé\" n'apparaissent pas dans la page Zabbix hostinventories.php, mais restent supervisés (alertes incluses)."
    ws2[f'A{note_row}'].font = Font(name='Calibri', size=9, italic=True, color='808080')
    ws2[f'A{note_row}'].alignment = Alignment(vertical='center', wrap_text=True)

    # Feuille 3 : Alertes filtrées
    ws3 = wb.create_sheet("Alertes filtrées"); ws3.sheet_properties.tabColor = "999999"
    for col, w in {'A': 16, 'B': 18, 'C': 12, 'D': 55, 'E': 22}.items(): ws3.column_dimensions[col].width = w
    ws3.merge_cells('A1:E1'); ws3['A1'] = f'ALERTES FILTRÉES ({excluded_count})'; ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color='999999')
    for ci, hdr in enumerate(['Date', 'Hôte', 'Sévérité', 'Problème', 'Raison'], 1):
        cell = ws3.cell(row=3, column=ci, value=hdr); cell.font = Font(name='Calibri', bold=True, size=10, color='666666'); cell.fill = PatternFill('solid', fgColor='EEEEEE'); cell.border = thin_border
    frow = 4
    for p in problems:
        if not is_excluded(p): continue
        ti = triggers_map.get(p.get("objectid", ""), {}); hosts_list = ti.get("hosts") or []; hn = hosts_list[0].get("name", "") if hosts_list else ""
        try: et = datetime.fromtimestamp(int(p.get("clock", 0)))
        except: et = now
        reason = "Sévérité Information" if str(p.get("severity", "0")) in EXCLUDED_SEVERITIES else "Pattern exclu"
        for ci, v in enumerate([et.strftime("%d/%m/%Y %H:%M"), hn, severity_name(p.get("severity", "0")), p.get("name", ""), reason], 1):
            cell = ws3.cell(row=frow, column=ci, value=v); cell.font = Font(name='Calibri', size=9, color='999999'); cell.border = thin_border
        frow += 1

    # Feuille 4 : Analyse
    ws4 = wb.create_sheet("Analyse"); ws4.sheet_properties.tabColor = "C0392B"
    for col, w in {'A': 22, 'B': 18, 'C': 55, 'D': 10, 'E': 20}.items(): ws4.column_dimensions[col].width = w
    ws4.merge_cells('A1:E1'); ws4['A1'] = 'ANALYSE DES PROBLÈMES'; ws4['A1'].font = title_font; ws4.row_dimensions[1].height = 35
    for c in range(1, 6): ws4.cell(row=2, column=c).border = bottom_border
    hors_reseau = []; problemes_reels = []
    by_subject = {"Espace disque": [], "Agent indisponible": [], "Service arrêté": [], "Mémoire / CPU": [], "Interface réseau": [], "Équipement injoignable": [], "Autre": []}
    for p in filtered:
        ti = triggers_map.get(p.get("objectid", ""), {}); hosts_list = ti.get("hosts") or []; hn = hosts_list[0].get("name", "") if hosts_list else ""
        try: et = datetime.fromtimestamp(int(p.get("clock", 0)))
        except: et = now
        days = (now - et).days; name = p.get("name", ""); sev = severity_name(p.get("severity", "0")); entry = {"host": hn, "problem": name, "days": days, "severity": sev}; name_lower = name.lower()
        if "space" in name_lower or "disque" in name_lower or "fs [" in name_lower: by_subject["Espace disque"].append(entry)
        elif "not available" in name_lower or "agent" in name_lower: by_subject["Agent indisponible"].append(entry)
        elif "service" in name_lower or "not running" in name_lower or "windefend" in name_lower: by_subject["Service arrêté"].append(entry)
        elif "memory" in name_lower or "cpu" in name_lower or "load" in name_lower: by_subject["Mémoire / CPU"].append(entry)
        elif "link down" in name_lower or "interface" in name_lower or "bandwidth" in name_lower: by_subject["Interface réseau"].append(entry)
        elif "unavailable by icmp" in name_lower or "unreachable" in name_lower: by_subject["Équipement injoignable"].append(entry)
        else: by_subject["Autre"].append(entry)
        if days > 30 and ("not available" in name_lower or "unavailable" in name_lower or "link down" in name_lower): hors_reseau.append(entry)
        elif days > 7 and ("space" in name_lower or "service" in name_lower or "memory" in name_lower or "not running" in name_lower or "toner" in name_lower): problemes_reels.append(entry)
    row = 4
    ws4.merge_cells(f'A{row}:E{row}'); ws4[f'A{row}'] = f'  ÉQUIPEMENTS PROBABLEMENT HORS RÉSEAU ({len(hors_reseau)})'; ws4[f'A{row}'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); ws4[f'A{row}'].fill = PatternFill('solid', fgColor='922B21'); ws4[f'A{row}'].alignment = Alignment(vertical='center'); ws4.row_dimensions[row].height = 28; row += 1
    if hors_reseau:
        for ci, hdr in enumerate(['Hôte', 'Sévérité', 'Problème', 'Jours', 'Statut'], 1):
            cell = ws4.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='C0392B'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
        row += 1
        for entry in sorted(hors_reseau, key=lambda x: -x["days"]):
            for ci, v in enumerate([entry["host"], entry["severity"], entry["problem"][:50], entry["days"], "À supprimer ?"], 1):
                cell = ws4.cell(row=row, column=ci, value=v); cell.font = normal_font; cell.fill = PatternFill('solid', fgColor='FADBD8'); cell.border = thin_border
            row += 1
    else: ws4[f'A{row}'] = '  Aucun équipement suspect détecté'; ws4[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='808080'); row += 1
    row += 1
    ws4.merge_cells(f'A{row}:E{row}'); ws4[f'A{row}'] = f'  PROBLÈMES RÉELS À TRAITER ({len(problemes_reels)})'; ws4[f'A{row}'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); ws4[f'A{row}'].fill = PatternFill('solid', fgColor='D4AC0D'); ws4[f'A{row}'].alignment = Alignment(vertical='center'); ws4.row_dimensions[row].height = 28; row += 1
    # Note : la priorité ci-dessous est basée sur l'ancienneté, PAS sur la sévérité Zabbix (page 1).
    ws4.merge_cells(f'A{row}:E{row}')
    ws4[f'A{row}'] = "  Priorité calculée sur l'ancienneté de l'alerte (URGENT > 60 j, IMPORTANT > 30 j, À PLANIFIER <= 30 j). Échelle distincte de la sévérité Zabbix affichée en page 1."
    ws4[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='808080'); ws4[f'A{row}'].alignment = Alignment(vertical='center', wrap_text=True); ws4.row_dimensions[row].height = 26; row += 1
    if problemes_reels:
        for ci, hdr in enumerate(['Hôte', 'Sévérité', 'Problème', 'Jours', 'Priorité (ancienneté)'], 1):
            cell = ws4.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='B7950B'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
        row += 1
        for entry in sorted(problemes_reels, key=lambda x: -x["days"]):
            prio = "URGENT" if entry["days"] > 60 else "IMPORTANT" if entry["days"] > 30 else "À PLANIFIER"
            row_fill = PatternFill('solid', fgColor='FEF9E7') if entry["days"] <= 30 else PatternFill('solid', fgColor='FDEBD0')
            for ci, v in enumerate([entry["host"], entry["severity"], entry["problem"][:50], entry["days"], prio], 1):
                cell = ws4.cell(row=row, column=ci, value=v); cell.font = normal_font; cell.fill = row_fill; cell.border = thin_border
            pc = ws4.cell(row=row, column=5)
            if prio == "URGENT": pc.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF'); pc.fill = PatternFill('solid', fgColor='C0392B')
            elif prio == "IMPORTANT": pc.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF'); pc.fill = PatternFill('solid', fgColor='E67E22')
            pc.alignment = Alignment(horizontal='center'); row += 1
    else: ws4[f'A{row}'] = '  Aucun problème réel détecté'; ws4[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='808080'); row += 1
    row += 1
    # Section : alertes résolues depuis le dernier rapport
    ws4.merge_cells(f'A{row}:E{row}'); ws4[f'A{row}'] = f'  ALERTES RÉSOLUES DEPUIS LE DERNIER RAPPORT ({len(resolved_list)})'; ws4[f'A{row}'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); ws4[f'A{row}'].fill = PatternFill('solid', fgColor='1E8449'); ws4[f'A{row}'].alignment = Alignment(vertical='center'); ws4.row_dimensions[row].height = 28; row += 1
    if resolved_list:
        ws4.merge_cells(f'A{row}:E{row}')
        ws4[f'A{row}'] = "  Le symbole # remplace les valeurs variables du libellé (tailles, %, durées)."
        ws4[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='808080'); row += 1
        for ci, hdr in enumerate(['Hôte', 'Alerte (motif)', '', 'Vu (sem.)', ''], 1):
            cell = ws4.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='27AE60'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
        ws4.merge_cells(f'B{row}:C{row}'); row += 1
        for entry in resolved_list[:20]:
            ws4.cell(row=row, column=1, value=entry["host"]).font = normal_font
            ws4.merge_cells(f'B{row}:C{row}')
            ws4.cell(row=row, column=2, value=entry["pattern"][:90]).font = normal_font
            ws4.cell(row=row, column=4, value=entry["count"]).font = normal_font
            ws4.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            for c in range(1, 6): ws4.cell(row=row, column=c).fill = PatternFill('solid', fgColor='E8F5E9'); ws4.cell(row=row, column=c).border = thin_border
            row += 1
    else:
        ws4[f'A{row}'] = '  Aucune alerte résolue depuis le dernier rapport'; ws4[f'A{row}'].font = Font(name='Calibri', size=10, italic=True, color='808080'); row += 1
    row += 1
    # Section : tendance hebdomadaire (historique du volume d'alertes)
    ws4.merge_cells(f'A{row}:E{row}'); ws4[f'A{row}'] = '  TENDANCE HEBDOMADAIRE (12 dernières semaines max)'; ws4[f'A{row}'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); ws4[f'A{row}'].fill = PatternFill('solid', fgColor='1F4E79'); ws4[f'A{row}'].alignment = Alignment(vertical='center'); ws4.row_dimensions[row].height = 28; row += 1
    for ci, hdr in enumerate(['Semaine', 'Date du rapport', 'Alertes', 'Critiques', ''], 1):
        cell = ws4.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='2C3E50'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
    row += 1
    for h_entry in history:
        crit = int(h_entry.get("critiques", 0) or 0)
        for ci, v in enumerate([h_entry.get("week", ""), h_entry.get("date", ""), h_entry.get("alertes", ""), crit, ""], 1):
            cell = ws4.cell(row=row, column=ci, value=v); cell.font = normal_font; cell.border = thin_border
            if ci in (3, 4): cell.alignment = Alignment(horizontal='center')
        if crit > 0:
            cc = ws4.cell(row=row, column=4); cc.font = Font(name='Calibri', size=10, bold=True, color='C62828')
        row += 1
    row += 1
    ws4.merge_cells(f'A{row}:E{row}'); ws4[f'A{row}'] = '  RÉPARTITION PAR SUJET'; ws4[f'A{row}'].font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); ws4[f'A{row}'].fill = PatternFill('solid', fgColor='1F4E79'); ws4[f'A{row}'].alignment = Alignment(vertical='center'); ws4.row_dimensions[row].height = 28; row += 1
    subject_colors = {"Espace disque": "E74C3C", "Agent indisponible": "E67E22", "Service arrêté": "8E44AD", "Mémoire / CPU": "2980B9", "Interface réseau": "27AE60", "Équipement injoignable": "C0392B", "Autre": "95A5A6"}
    for ci, hdr in enumerate(['Sujet', 'Nb alertes', 'Hôtes concernés', '', ''], 1):
        cell = ws4.cell(row=row, column=ci, value=hdr); cell.font = header_font; cell.fill = PatternFill('solid', fgColor='2C3E50'); cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border
    row += 1
    for subject, items in by_subject.items():
        if not items: continue
        hosts_list_str = ", ".join(sorted(set(e["host"] for e in items)))[:60]; color = subject_colors.get(subject, "95A5A6")
        ws4.cell(row=row, column=1, value=subject).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF'); ws4.cell(row=row, column=1).fill = PatternFill('solid', fgColor=color)
        ws4.cell(row=row, column=2, value=len(items)).font = Font(name='Calibri', size=12, bold=True); ws4.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws4.merge_cells(f'C{row}:E{row}'); ws4.cell(row=row, column=3, value=hosts_list_str).font = Font(name='Calibri', size=9)
        for c in range(1, 6): ws4.cell(row=row, column=c).border = thin_border
        row += 1
    # Feuille 5 : Fonctionnement (explication non technique des recommandations)
    ws5 = wb.create_sheet("Fonctionnement"); ws5.sheet_properties.tabColor = "6A1B9A"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']: ws5.column_dimensions[col].width = 20
    r5 = [2]
    def w5(text, kind="text"):
        row_i = r5[0]
        ws5.merge_cells(f'A{row_i}:F{row_i}')
        c = ws5[f'A{row_i}']
        c.value = text
        if kind == "title":
            c.font = title_font; c.alignment = Alignment(vertical='center'); ws5.row_dimensions[row_i].height = 35
        elif kind == "section":
            c.value = f'  {text}'; c.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='6A1B9A'); c.alignment = Alignment(vertical='center'); ws5.row_dimensions[row_i].height = 26
        elif kind == "blank":
            ws5.row_dimensions[row_i].height = 8
        else:
            c.font = Font(name='Calibri', size=10); c.alignment = Alignment(vertical='top', wrap_text=True)
            nb = max(1, -(-len(text) // 110)); ws5.row_dimensions[row_i].height = nb * 14 + 6
        r5[0] += 1
    w5("COMMENT SONT PRODUITES LES RECOMMANDATIONS", "title")
    w5("La colonne « Recommandation » du rapport hebdomadaire est remplie automatiquement. Chaque recommandation provient de l'une des deux sources ci-dessous, toujours dans cet ordre.")
    w5("", "blank")
    w5("1. LE RÉFÉRENTIEL DÉTERMINISTE (source principale)", "section")
    w5("C'est une bibliothèque de fiches de réponses rédigées à l'avance et validées par l'équipe IT. Chaque fiche correspond à une famille d'alertes connue : disque plein, service arrêté, agent de supervision injoignable, machine redémarrée, forte charge mémoire ou CPU, horloge désynchronisée, port réseau tombé, consommable d'imprimante, etc.")
    w5("Concrètement, à la génération du rapport, le script fait trois choses pour chaque alerte :")
    w5("   1) Il lit le libellé de l'alerte et y cherche des mots-clés (par exemple « space is low » = disque plein).")
    w5("   2) Il identifie le type de machine concernée (Windows, Linux ou switch réseau) afin de proposer les commandes adaptées à ce système.")
    w5("   3) Il insère la fiche correspondante en la complétant avec les vraies valeurs de l'alerte : nom du service en panne, lettre du disque plein, numéro du port réseau.")
    w5("« Déterministe » signifie que la même alerte produit toujours exactement la même recommandation. Il n'y a aucune improvisation : ce sont des procédures écrites et relues par un humain. L'intelligence artificielle n'intervient pas du tout à cette étape.")
    w5("", "blank")
    w5("2. L'IA LOCALE (utilisée uniquement pour les cas inconnus)", "section")
    w5("L'IA intervient à un seul moment précis : lorsqu'une alerte ne correspond à aucune fiche du référentiel, c'est-à-dire un cas que l'équipe n'a pas encore documenté.")
    w5("Dans ce cas, le script envoie le libellé de l'alerte à une IA (modèle Qwen2.5) installée sur nos propres serveurs. Aucune donnée ne quitte l'entreprise : ce n'est pas un service en ligne comme ChatGPT, tout se passe en interne.")
    w5("L'IA reçoit des consignes strictes : réponse de 2 phrases maximum, uniquement des commandes de diagnostic (jamais de commande qui modifie ou supprime quoi que ce soit), et le type de machine lui est imposé par le script (elle ne le devine pas, pour éviter de proposer une commande Windows sur un serveur Linux).")
    w5("Sa réponse est ensuite nettoyée et raccourcie avant d'être insérée dans le rapport.")
    w5("", "blank")
    w5("3. COMMENT SAVOIR QUI A RÉPONDU ?", "section")
    w5("Lors de chaque génération, le script indique la source utilisée pour chaque alerte et affiche un résumé final, par exemple : « 12 via référentiel déterministe, 1 via IA locale ».")
    w5("En pratique, la quasi-totalité des alertes de notre parc correspond à des cas connus : le référentiel répond presque toujours, l'IA ne traite que l'inédit. À mesure que de nouveaux cas sont documentés, ils sont ajoutés au référentiel et l'IA intervient de moins en moins.")
    w5("", "blank")
    w5("4. PETIT LEXIQUE DU RAPPORT", "section")
    w5("Vu depuis : NOUVELLE = alerte apparue cette semaine ; « 5 sem. » = alerte présente depuis 5 rapports hebdomadaires consécutifs.")
    w5("Ack (acquittée) : « Oui » = un technicien a signalé dans Zabbix qu'il a vu l'alerte et la prend en charge ; « Non » = personne ne s'en occupe encore.")
    w5("Évolution (page 1) : nombre d'alertes nouvelles cette semaine, persistantes, et résolues depuis le rapport précédent.")
    w5("Alertes résolues (onglet Analyse) : alertes présentes dans le rapport précédent et qui ont disparu depuis.")
    w5("Tendance hebdomadaire (onglet Analyse) : volume total d'alertes semaine par semaine, pour voir si la situation s'améliore ou se dégrade.")
    return wb

def send_email(filepath, filename):
    msg = MIMEMultipart()
    msg['From'] = SMTP_FROM
    msg['To'] = ", ".join(EMAIL_TO)
    msg['Subject'] = f"Rapport Zabbix hebdomadaire - {datetime.now().strftime('%d/%m/%Y')}"
    msg.attach(MIMEText(f"Bonjour,\n\nVeuillez trouver ci-joint le rapport Zabbix de la semaine.\nRecommandations : référentiel déterministe + IA locale Qwen2.5 (explications dans l'onglet Fonctionnement du fichier).\n\nGénéré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}.\n\n---\nRapport généré automatiquement", 'plain', 'utf-8'))
    with open(filepath, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read()); encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"'); msg.attach(part)
    try:
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_SENDER, EMAIL_TO, msg.as_string())
        print(f"[OK] Email envoyé à {', '.join(EMAIL_TO)}")
    except Exception as e:
        print(f"[ERREUR] Email: {e}")

def build_arg_parser():
    """Construit le parser argparse (sans effet de bord, testable)."""
    p = argparse.ArgumentParser(
        prog="zabbix_auto_report",
        description="Génère le rapport Zabbix hebdomadaire Excel (référentiel déterministe "
                    "+ IA locale Ollama pour les cas inconnus). Aucune donnée ne quitte le serveur.",
        epilog="Configuration : variables d'environnement ou fichier .env (voir .env.example).",
    )
    p.add_argument("--no-email", action="store_true",
                   help="génère le rapport sans envoyer d'email (utile pour les tests)")
    p.add_argument("--test-email", action="store_true",
                   help="envoie un email de test via la configuration SMTP puis s'arrête "
                        "(aucun appel Zabbix / Ollama)")
    return p


def main(argv=None):
    args = build_arg_parser().parse_args(argv)

    os.makedirs(REPORT_DIR, exist_ok=True)

    # --test-email : envoi d'un email de test sans aucune connexion sortante Zabbix.
    if args.test_email:
        print("Test d'envoi d'email...")
        if not SMTP_SERVER or not SMTP_USER:
            print("[ERREUR] SMTP non configuré (SMTP_SERVER/SMTP_USER vides).")
            return 1
        msg = MIMEMultipart(); msg['From'] = SMTP_FROM; msg['To'] = ", ".join(EMAIL_TO); msg['Subject'] = "Test - Rapport Zabbix (IA)"
        msg.attach(MIMEText("Test avec recommandations IA. Configuration OK.", 'plain', 'utf-8'))
        try:
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
                server.login(SMTP_USER, SMTP_PASS); server.sendmail(SMTP_SENDER, EMAIL_TO, msg.as_string())
            print("[OK] Email de test envoyé !")
            return 0
        except Exception as e:
            print(f"[ERREUR] {e}")
            return 1

    print("Vérification Ollama...", end=" ", flush=True)
    try:
        test = ask_ollama("Reponds OK")
        if test: print("[OK] Ollama opérationnel")
        else: print("[WARN] Ollama ne répond pas")
    except: print("[WARN] Ollama indisponible")
    print("Connexion à l'API Zabbix..."); auth = get_auth_token(); print("[OK] Connecté")
    print(f"Version API Zabbix : {get_api_version(auth)}")
    print("Récupération des hôtes..."); hosts = get_hosts(auth); print(f"[OK] {len(hosts)} hôtes")
    print("Récupération de la disponibilité..."); availability = get_host_availability(auth); print(f"[OK] {availability['available']} dispo / {availability['unavailable']} down / {availability['unknown']} inconnus")
    print("Récupération des problèmes..."); problems = get_problems(auth); print(f"[OK] {len(problems)} problèmes bruts")
    trigger_ids = list({p.get("objectid") for p in problems if p.get("objectid")})
    print("Récupération des triggers..."); triggers = get_triggers(auth, trigger_ids); triggers_map = {t["triggerid"]: t for t in triggers}; print(f"[OK] {len(triggers)} triggers")
    try: zabbix_api("user.logout", [], auth)
    except: pass
    filtered = [p for p in problems if not is_excluded(p)]
    print(f"[OK] {len(filtered)} alertes pertinentes ({len(problems)-len(filtered)} filtrées)")
    print(f"Génération des recommandations ({len(filtered)} alertes)...")
    wb = generate_report(hosts, problems, triggers_map, availability)
    filename = f"rapport_zabbix_{datetime.now().strftime('%Y-%m-%d')}.xlsx"; filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath); print(f"[OK] Rapport: {filepath}")
    if not args.no_email:
        print("Envoi par email..."); send_email(filepath, filename)
    print("\nTermine !")
    return 0


if __name__ == "__main__":
    sys.exit(main())
